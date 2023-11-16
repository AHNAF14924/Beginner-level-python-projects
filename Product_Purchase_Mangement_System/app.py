import openpyxl
from tabulate import tabulate
import sys
import pdfkit

# Load Excel file
file_path = 'products_list.xlsx'
workbook = openpyxl.load_workbook(file_path)


def load_sheet_options():
    return {index + 1: sheet_name for index, sheet_name in enumerate(workbook.sheetnames)}


def display_product_options():
    print("Please Choose an Options:")
    for option, sheet_name in sheet_options.items():
        print(f"{option}. {sheet_name}")
    print(f"{len(sheet_options) + 1}. Checkout")


def process_product_choice(product_category):
    sheet = workbook[product_category]
    header = [cell.value for cell in sheet[1]]
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(list(row))

    print(f"\n{product_category} :")
    numbered_data = [[index + 1] + row for index, row in enumerate(data)]
    print(tabulate(numbered_data, headers=[
          ''] + header, tablefmt='grid', showindex=False))

    product_choice = int(
        input("Enter the number of the product you want to buy: "))
    selected_product = data[product_choice - 1]
    brand_index, model_index, price_index = header.index(
        'Brand'), header.index('Model'), header.index('Price')
    selected_product_name = f"{selected_product[brand_index]} {selected_product[model_index]}"
    print(f"You have selected: {selected_product_name}")

    quantity = int(input("Enter the quantity you want to buy: "))
    unit_price = selected_product[price_index]
    total_price = unit_price * quantity

    return {
        "Product Name": selected_product_name,
        "Qty": quantity,
        "Unit Price": unit_price,
        "Total Price": total_price
    }


def checkout(purchase_history):
    if not purchase_history:
        print("Your cart is empty. Please add items to your cart before checkout.")
        return purchase_history

    grand_total = sum(int(item['Total Price']) for item in purchase_history)
    invoice_table = []
    for item in purchase_history:
        invoice_table.append([
            item['Product Name'],
            item['Qty'],
            item['Unit Price'],
            item['Total Price']
        ])

    invoice_table.append(["Grand Total", "", "", grand_total])

    html_content = tabulate(invoice_table, headers=[
                            'Product Name', 'Qty', 'Unit Price', 'Total Price'], tablefmt='html')
    pdfkit.from_string(html_content, 'invoice.pdf')

    return []


def main():
    global sheet_options
    sheet_options = load_sheet_options()

    print("Welcome to Our Store!")

    purchase_history = []  # Initialize purchase_history here

    while True:
        display_product_options()
        choice = int(input("Select an option: "))

        if choice == len(sheet_options) + 1:
            purchase_history = checkout(purchase_history)
            print("Thank you for shopping with us!")
            sys.exit()
        elif choice in sheet_options:
            product_info = process_product_choice(sheet_options[choice])
            purchase_history.append(product_info)
            print(
                f"Added '{product_info['Product Name']}' to your cart. Total price: {product_info['Total Price']}")
        else:
            print("Invalid option. Please select a valid option.")


if __name__ == "__main__":
    main()
